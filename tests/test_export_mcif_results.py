import importlib.util
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parent.parent


def _load_export_mcif_results_module():
    module_path = PROJECT_ROOT / "scripts" / "export_mcif_results_to_excel.py"
    spec = importlib.util.spec_from_file_location("export_mcif_results_to_excel", module_path)
    assert spec is not None
    assert spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def test_runtime_excel_export_rows_include_quasi2d_fields():
    exporter = _load_export_mcif_results_module()
    row = exporter._row_from_serialized_result_record(
        {
            "case_id": "quasi.mcif",
            "file_name": "quasi.mcif",
            "status": "ok",
            "duration_seconds": 1.25,
            "result": {
                "index": "123.47.1.1.L",
                "conf": "Collinear",
                "phase": "AFM",
                "acc": "4/mmmP",
                "convention_spin_only_direction": "0,0,1",
                "convention_ssg_international_linear": "P 1",
                "convention_ssg_international_latex": "P 1",
                "gspg_effective_mpg_symbol": "mmm1'",
                "spinsplitting_w_soc": "spin splitting",
                "spinsplitting_wo_soc": "spin splitting",
                "spin_texture_config_no_soc": {
                    "spin_texture_type": "d-wave",
                    "momentum_space_spin_configuration": "coplanar",
                },
                "spin_texture_config_soc": {
                    "spin_texture_type": "s-wave",
                    "momentum_space_spin_configuration": "noncoplanar",
                },
                "ahc_w_soc": "No",
                "ahc_wo_soc": "No",
                "tensor_outputs": {
                    "AHE_wSOC": {
                        "is_zero": False,
                        "relations": [r"\sigma_{xy} = -\sigma_{yx}"],
                        "components": [["0", "C1", "0"], ["-C1", "0", "0"], ["0", "0", "0"]],
                    },
                    "AHE_woSOC": {"is_zero": True},
                    "QMDTensor": {"is_zero": True},
                    "MSGQMDTensor": {
                        "is_zero": False,
                        "relations": [],
                        "components": [
                            [["C1", "0", "0"], ["0", "0", "0"], ["0", "0", "0"]],
                            [["0", "0", "0"], ["0", "0", "0"], ["0", "0", "0"]],
                            [["0", "0", "0"], ["0", "0", "0"], ["0", "0", "0"]],
                        ],
                    },
                    "IMDTensor": {"is_zero": True},
                    "MSGIMDTensor": {"is_zero": True},
                    "BCDTensor": {"is_zero": True},
                    "MSGBCDTensor": {"is_zero": True},
                },
                "is_alter": "(Altermagnet)",
                "is_spin_orbit_magnet": "",
                "polar_axes_by_symmetry": {
                    "sg": {
                        "allowed_polar_axes": [{"label": "z", "setting": "G0std"}],
                        "allowed_polar_axes_setting": "G0std",
                    },
                    "ossg": {
                        "allowed_polar_axes": [{"label": "x", "setting": "G0std"}],
                        "allowed_polar_axes_setting": "G0std",
                    },
                    "msg": {
                        "allowed_polar_axes": [],
                        "allowed_polar_axes_setting": "G0std",
                    },
                },
                "wp_chain": [
                    ["Fe", "2a", 0, "1a", 0, "1a", 0],
                    ["Fe", "2a", 0, "1b", 1, "1b", 1],
                    ["Mn", "1c", 1, "1c", 2, "1d", 2],
                    ["Mn", "1c", 1, "1c", 2, "1e", 3],
                ],
                "magnetic_site_summary": {
                    "status": "ok",
                    "setting": "G0_standard",
                    "SG": {"number": 164, "symbol": "P-3m1"},
                    "cell_expansion": 4,
                    "ssg_index": "12.2.2.1.P3",
                    "magnetic_atom_count": 8,
                    "nonzero_moment_atom_count": 6,
                    "zero_moment_magnetic_atom_count": 2,
                    "magnetic_atom_selection_mode": "sg_orbit_closure_of_nonzero_moment_sites",
                    "n_magnetic_orbits_sg": 1,
                    "n_magnetic_orbits_ssg": 1,
                    "n_magnetic_orbits_msg": 1,
                    "max_magnetic_site_dof_ssg": 2,
                    "max_magnetic_site_dof_msg": 3,
                    "total_magnetic_site_dof_ssg": 2,
                    "total_magnetic_site_dof_msg": 3,
                    "magnetic_wp_dof_rows": [
                        {
                            "element": "Fe",
                            "site_count": 4,
                            "site_indices": [0, 1, 2, 3],
                            "sg_wyckoff": "12f",
                            "sg_wyckoff_index": 1,
                            "ssg_wyckoff": "4d",
                            "ssg_wyckoff_with_dof": "4d(2)",
                            "ssg_wyckoff_index": 3,
                            "ssg_site_dof": 2,
                            "ssg_orbit_total_dof": 2,
                            "ssg_constraints": ["Sx", "Sy", "0"],
                            "ssg_representative_index": 0,
                            "msg_wyckoff": "4d",
                            "msg_wyckoff_with_dof": "4d(3)",
                            "msg_wyckoff_index": 3,
                            "msg_site_dof": 3,
                            "msg_orbit_total_dof": 3,
                            "msg_constraints": ["Sx", "Sy", "Sz"],
                            "msg_representative_index": 0,
                        }
                    ],
                },
                "quasi_2d": {
                    "calculation_mode": "quasi2d",
                    "dimension": "2d",
                    "status": "heuristic",
                    "source": "heuristic",
                    "vacuum_axis_input": "c",
                    "spin_splitting_2d": "spin splitting",
                    "interpretation": "in_plane_k_dependent",
                    "is_alter_2d": "(Altermagnet)",
                    "magnetic_phase": "AFM(Altermagnet)",
                    "diagnostic_points": [
                        {
                            "label": "GP",
                            "k_symbol_2d": "GP:(0.237,0.371,0)",
                            "k_input_reciprocal": [0.237, 0.371, 0.0],
                            "k_acc_primitive": [0.237, 0.371, 0.0],
                            "spin_splitting": "spin splitting",
                            "spin_polarizations": ["Sx", "0", "0"],
                        }
                    ],
                    "kpoints": [
                        {
                            "label": "Γ",
                            "k_symbol_2d": "Γ:(0,0,0)",
                            "plane_classification": "in_plane",
                            "spin_splitting": "no spin splitting",
                        },
                    ],
                    "kpoint_projection_summary": {
                        "source": "acc_table",
                        "total": 2,
                        "by_plane_count": {"in_plane": 1, "mixed": 1},
                    },
                    "generic_point_comparison": {
                        "status": "compared",
                        "gp_3d": {
                            "label": "GP",
                            "k_symbol_2d": "GP:(0.417,0.237,0.371)",
                            "k_input_reciprocal": [0.417, 0.237, 0.371],
                            "k_acc_primitive": [0.417, 0.237, 0.371],
                            "plane_classification": "mixed",
                            "spin_splitting": "no spin splitting",
                            "spin_polarizations": [],
                        },
                        "gp_2d": {
                            "label": "GP",
                            "k_symbol_2d": "GP:(0.237,0.371,0)",
                            "k_input_reciprocal": [0.237, 0.371, 0.0],
                            "k_acc_primitive": [0.237, 0.371, 0.0],
                            "plane_classification": "in_plane",
                            "spin_splitting": "spin splitting",
                            "spin_polarizations": ["Sx", "0", "0"],
                        },
                        "k_input_delta_wrapped": [-0.18, 0.134, -0.371],
                        "k_acc_delta_wrapped": [-0.18, 0.134, -0.371],
                        "k_input_changed": True,
                        "spin_splitting_changed": True,
                        "spin_polarization_changed": True,
                        "summary": "k_changed_spin_splitting_changed",
                    },
                    "spin_texture_config_no_soc": {
                        "spin_texture_type": "d-wave",
                        "basis_setting": "quasi2d_ossg_unit_cartesian_in_plane",
                    },
                    "spin_texture_config_soc": {
                        "spin_texture_type": "s-wave",
                        "basis_setting": "quasi2d_ossg_unit_cartesian_in_plane",
                    },
                    "spin_texture_config_basis": {
                        "setting": "quasi2d_ossg_unit_cartesian_in_plane",
                        "in_plane_k_axes": ["a", "b"],
                    },
                },
            },
        }
    )

    assert "calculation_mode" not in row
    assert "dimension" not in row
    assert row["quasi2d_status"] == "heuristic"
    assert row["quasi2d_source"] == "heuristic"
    assert row["vacuum_axis_input"] == "c"
    assert row["spin_splitting_2d"] == "spin splitting"
    assert row["spin_splitting_2d_interpretation"] == "in_plane_k_dependent"
    assert row["is_alter_2d"] == "(Altermagnet)"
    assert row["quasi2d_magnetic_phase"] == "AFM(Altermagnet)"
    assert row["quasi2d_gp_label"] == "GP"
    assert row["quasi2d_gp_symbol"] == "GP:(0.237,0.371,0)"
    assert row["quasi2d_gp_spin_splitting"] == "spin splitting"
    assert row["quasi2d_spin_texture_config_no_soc"]["spin_texture_type"] == "d-wave"
    assert row["quasi2d_spin_texture_config_soc"]["spin_texture_type"] == "s-wave"
    assert row["quasi2d_spin_texture_basis"]["in_plane_k_axes"] == ["a", "b"]
    assert "quasi2d_3d_gp_symbol" not in row
    assert "quasi2d_2d_gp_symbol" not in row
    assert "quasi2d_gp_spin_splitting_changed" not in row
    assert row["quasi2d_kpoint_projection_summary"]["by_plane_count"]["mixed"] == 1
    assert row["quasi2d_kpoints"] == [
        {
            "label": "Γ",
            "k_symbol_2d": "Γ:(0,0,0)",
            "plane": "in_plane",
            "spin_splitting": "no spin splitting",
        }
    ]
    assert row["spin_splitting_with_soc"] == "spin splitting"
    assert row["spin_splitting_without_soc"] == "spin splitting"
    assert row["spin_only_direction(ossg convention)"] == "0,0,1"
    assert row["ossg_symbol_linear"] == "P 1"
    assert row["ossg_symbol_latex"] == "P 1"
    assert row["empg_symbol"] == "mmm1'"
    assert row["spin_texture_config_no_soc"]["spin_texture_type"] == "d-wave"
    assert row["spin_texture_type_no_soc"] == "d-wave"
    assert row["momentum_space_spin_configuration_no_soc"] == "coplanar"
    assert row["spin_texture_type_soc"] == "s-wave"
    assert row["momentum_space_spin_configuration_soc"] == "noncoplanar"
    assert row["ahc_with_soc"] == "No"
    assert row["ahc_without_soc"] == "No"
    assert r"\sigma_{xy} = -\sigma_{yx}" in row["ahe_tensor_equations_soc"]
    assert r"\sigma_{xy} = C1" in row["ahe_tensor_equations_soc"]
    assert row["ahe_tensor_equations_no_soc"] == r"\sigma=0"
    assert "Q_{xxx} = C1" in row["qmd_tensor_equations_soc"]
    assert row["sg_polar_axes"] == "z"
    assert row["sg_polar_axes_setting"] == "G0std"
    assert row["ossg_polar_axes"] == "x"
    assert row["msg_polar_axes"] is None
    assert row["has_wyckoff_splitting_sg_to_ossg"] is True
    assert row["has_wyckoff_splitting_ossg_to_msg"] is True
    assert row["is_altermagnet"] == "(Altermagnet)"
    assert row["is_spin_orbit_magnet"] == ""
    assert "status" not in row
    assert "magnetic_site_status" not in row
    assert row["magnetic_site_sg_primitive_to_magnetic_primitive_cell_expansion"] == 4
    assert row["magnetic_atom_count"] == 8
    assert row["nonzero_moment_atom_count"] == 6
    assert row["zero_moment_magnetic_atom_count"] == 2
    assert "magnetic_atom_selection_mode" not in row
    assert "magnetic_site_cell_expansion" not in row
    assert "magnetic_site_nonmagnetic_sg_num" not in row
    assert "magnetic_site_nonmagnetic_sg_symbol" not in row
    assert "magnetic_site_ssg_index" not in row
    assert row["number_of_magnetic_orbits_sg"] == 1
    assert row["max_magnetic_site_dof_ssg"] == 2
    assert row["total_magnetic_site_dof_ssg"] == 2
    assert row["total_magnetic_site_dof_msg"] == 3
    assert row["magnetic_wyckoff_dof_summary"] == "Fe:12f->4d(2)->4d(3) n=4"
    assert "_magnetic_site_orbit_rows" in row
    assert row["_magnetic_site_orbit_rows"][0]["ssg_site_dof"] == 2


def test_runtime_excel_export_accepts_legacy_centrosymmetry_names():
    exporter = _load_export_mcif_results_module()

    row = exporter._row_from_serialized_result_record(
        {
            "case_id": "legacy.mcif",
            "file_name": "legacy.mcif",
            "status": "ok",
            "result": {
                "sg_has_real_space_inversion": True,
                "ossg_has_real_space_inversion": False,
                "msg_has_real_space_inversion": True,
            },
        }
    )

    assert row["sg_is_centrosymmetric"] is True
    assert row["ossg_is_centrosymmetric"] is False
    assert row["msg_is_centrosymmetric"] is True
    assert "sg_has_real_space_inversion" not in row


def test_excel_export_schema_is_shared_and_error_rows_are_complete(tmp_path):
    exporter = _load_export_mcif_results_module()

    assert exporter.COLUMNS == list(exporter.EXPORT_ROW_COLUMNS)
    assert exporter.MAGNETIC_ORBIT_COLUMNS == list(exporter.MAGNETIC_ORBIT_EXPORT_COLUMNS)
    assert exporter.QUASI2D_COLUMNS == list(exporter.QUASI2D_EXPORT_COLUMNS)
    assert len(exporter.COLUMNS) == len(set(exporter.COLUMNS))
    assert len(exporter.MAGNETIC_ORBIT_COLUMNS) == len(set(exporter.MAGNETIC_ORBIT_COLUMNS))
    assert "calculation_mode" not in exporter.COLUMNS
    assert "dimension" not in exporter.COLUMNS
    assert "source_route" not in exporter.COLUMNS
    assert "status" not in exporter.COLUMNS
    assert "magnetic_site_status" not in exporter.COLUMNS
    assert "magnetic_atom_selection_mode" not in exporter.COLUMNS
    assert "primitive_ssg_symbol" not in exporter.COLUMNS
    assert "wave" + "_spin_config_no_soc" not in exporter.COLUMNS
    assert "spin_texture_config_no_soc" in exporter.COLUMNS
    assert "spin_only_direction(ossg convention)" in exporter.COLUMNS
    assert "quasi2d_3d_gp_symbol" not in exporter.COLUMNS
    assert "quasi2d_status" not in exporter.COLUMNS
    assert "quasi2d_status" in exporter.QUASI2D_COLUMNS

    row = exporter._row_from_error(tmp_path / "broken.mcif", RuntimeError("boom"))

    for column in exporter.COLUMNS:
        assert column in row
    assert "status" not in row
    assert row["error_type"] == "RuntimeError"
    assert row["error_message"] == "boom"
    assert "quasi2d_magnetic_phase" not in row
    assert "quasi2d_gp_symbol" not in row


def test_workbook_appends_quasi2d_fields_to_records_only_when_present(tmp_path):
    exporter = _load_export_mcif_results_module()

    base_row = exporter.complete_export_row(
        {
            "case_id": "bulk",
            "file_name": "bulk.mcif",
            "status": "ok",
            "index": "1.1.1.1.P1",
        }
    )
    bulk_xlsx = tmp_path / "bulk.xlsx"
    exporter._write_workbook([base_row], bulk_xlsx)
    bulk_wb = load_workbook(bulk_xlsx)

    assert "records" in bulk_wb.sheetnames
    assert "magnetic_site_orbits" in bulk_wb.sheetnames
    assert "quasi2d" not in bulk_wb.sheetnames
    records_header = [cell.value for cell in bulk_wb["records"][1]]
    assert "quasi2d_status" not in records_header
    assert "calculation_mode" not in records_header

    quasi_row = dict(base_row)
    quasi_row.update(
        {
            "case_id": "layer",
            "file_name": "layer.mcif",
            "quasi2d_status": "heuristic",
            "quasi2d_gp_symbol": "GP:(u,v,0)",
        }
    )
    quasi_xlsx = tmp_path / "quasi.xlsx"
    exporter._write_workbook([quasi_row], quasi_xlsx)
    quasi_wb = load_workbook(quasi_xlsx)

    assert "quasi2d" not in quasi_wb.sheetnames
    quasi_header = [cell.value for cell in quasi_wb["records"][1]]
    assert "index" in quasi_header
    assert quasi_header.index("quasi2d_status") > quasi_header.index("error_message")
    assert "quasi2d_gp_symbol" in quasi_header
    assert "quasi2d_3d_gp_symbol" not in quasi_header


def test_workbook_writes_dotted_group_numbers_as_text(tmp_path):
    exporter = _load_export_mcif_results_module()

    row = exporter.complete_export_row(
        {
            "case_id": "msg",
            "file_name": "msg.mcif",
            "status": "ok",
            "index": "12.2.2.1.P3",
            "msg_num": 1234,
            "msg_bns_number": 12.34,
            "msg_og_number": "12.34.5",
            "_magnetic_site_orbit_rows": [
                {
                    "element": "Fe",
                    "site_count": 1,
                    "site_indices": [0],
                    "sg_wyckoff": "1a",
                    "sg_wyckoff_index": 0,
                    "ssg_wyckoff": "1a",
                    "ssg_wyckoff_with_dof": "1a(3)",
                    "ssg_wyckoff_index": 0,
                    "ssg_site_dof": 3,
                    "ssg_orbit_total_dof": 3,
                    "msg_wyckoff": "1a",
                    "msg_wyckoff_with_dof": "1a(3)",
                    "msg_wyckoff_index": 0,
                    "msg_site_dof": 3,
                    "msg_orbit_total_dof": 3,
                }
            ],
        }
    )
    xlsx = tmp_path / "msg_text.xlsx"
    exporter._write_workbook([row], xlsx)

    wb = load_workbook(xlsx)
    records = wb["records"]
    records_header = [cell.value for cell in records[1]]
    for column, expected in {
        "index": "12.2.2.1.P3",
        "msg_num": "1234",
        "msg_bns_number": "12.34",
        "msg_og_number": "12.34.5",
    }.items():
        cell = records.cell(row=2, column=records_header.index(column) + 1)
        assert cell.value == expected
        assert cell.number_format == "@"

    orbit_ws = wb["magnetic_site_orbits"]
    orbit_header = [cell.value for cell in orbit_ws[1]]
    orbit_index_cell = orbit_ws.cell(row=2, column=orbit_header.index("index") + 1)
    assert orbit_index_cell.value == "12.2.2.1.P3"
    assert orbit_index_cell.number_format == "@"


def test_export_can_read_parallel_shard_root(tmp_path):
    exporter = _load_export_mcif_results_module()
    root = tmp_path / "parallel"
    shard_1 = root / "shard_1"
    shard_0 = root / "shard_0"
    shard_1.mkdir(parents=True)
    shard_0.mkdir(parents=True)
    for shard in (shard_0, shard_1):
        (shard / "summary.json").write_text(
            '{"package_version":"0.test","route":"full"}',
            encoding="utf-8",
        )

    record_b = {
        "case_id": "cases/b.mcif",
        "file_name": "b.mcif",
        "status": "ok",
        "result": {"index": "2.2.2.2.P1"},
    }
    record_a = {
        "case_id": "cases/a.mcif",
        "file_name": "a.mcif",
        "status": "ok",
        "result": {"index": "1.1.1.1.P1"},
    }
    (shard_1 / "full_results.jsonl").write_text(
        exporter.json.dumps(record_b) + "\n",
        encoding="utf-8",
    )
    (shard_0 / "full_results.jsonl").write_text(
        exporter.json.dumps(record_a) + "\n",
        encoding="utf-8",
    )

    runtime_paths = exporter._discover_runtime_jsonls(
        None,
        shard_root=root,
        shard_glob="shard_*",
    )
    rows = exporter._rows_from_runtime_jsonls(runtime_paths, sort_rows=True)
    metadata = exporter._runtime_export_metadata_from_paths(runtime_paths)

    assert [path.parent.name for path in runtime_paths] == ["shard_0", "shard_1"]
    assert [row["case_id"] for row in rows] == ["cases/a.mcif", "cases/b.mcif"]
    assert metadata["source_run_tag"] == "parallel"
    assert metadata["source_fsg_version"] == "0.test"
