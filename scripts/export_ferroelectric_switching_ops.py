from __future__ import annotations

import argparse
import csv
import json
from collections import Counter
from pathlib import Path
from typing import Any


ROW_COLUMNS = [
    "case_id",
    "file_name",
    "index",
    "conf",
    "phase",
    "sg_num",
    "ossg_space_group_number",
    "msg_parent_space_group_number",
    "msg_symbol",
    "relation_layer",
    "screening_status",
    "basis_setting",
    "ordered_subgroup_source",
    "ordered_time_branch_scope",
    "parent_space_group_number",
    "ordered_space_group_number",
    "parent_grey_operation_count",
    "ordered_operation_count",
    "left_coset_count",
    "coset_index",
    "scope",
    "xyzt",
    "uvw",
    "coset_operation",
    "reverses_S",
    "reverses_P",
    "S_relation",
    "P_relation",
    "representative_class",
    "soc_allowed",
    "exchange_only",
    "magnetic_order_relation",
    "pattern_status",
]


def _get_nested(payload: dict[str, Any], path: str) -> Any:
    value: Any = payload
    for key in path.split("."):
        if not isinstance(value, dict):
            return None
        value = value.get(key)
    return value


def _result_payload(record: dict[str, Any]) -> dict[str, Any]:
    result = record.get("result")
    return result if isinstance(result, dict) else {}


def _case_value(record: dict[str, Any], result: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        if "." in key:
            value = _get_nested(result, key)
        else:
            value = result.get(key)
        if value is not None:
            return value
    for key in keys:
        if "." not in key:
            value = record.get(key)
            if value is not None:
                return value
    return None


def _screening_payload(ferroelectric: dict[str, Any], layer: str) -> dict[str, Any]:
    key = (
        "domain_reversal_symmetry_screening"
        if layer == "exchange_spin_space"
        else "soc_domain_reversal_symmetry_screening"
    )
    value = ferroelectric.get(key)
    return value if isinstance(value, dict) else {}


def _relation_rows(ferroelectric: dict[str, Any], layer: str) -> list[dict[str, Any]]:
    key = "domain_relation_rows" if layer == "exchange_spin_space" else "soc_domain_relation_rows"
    value = ferroelectric.get(key)
    return [row for row in value if isinstance(row, dict)] if isinstance(value, list) else []


def iter_switching_operation_rows(
    runtime_jsonl: Path,
    *,
    include_non_switching: bool = False,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with runtime_jsonl.open(encoding="utf-8") as handle:
        for line in handle:
            if not line.strip():
                continue
            record = json.loads(line)
            if record.get("status") != "ok":
                continue
            result = _result_payload(record)
            ferroelectric = result.get("ferroelectric_switching")
            if not isinstance(ferroelectric, dict):
                ferroelectric = _get_nested(result, "summary.ferroelectric_switching")
            if not isinstance(ferroelectric, dict):
                continue
            for layer in ("exchange_spin_space", "soc_magnetic"):
                screening = _screening_payload(ferroelectric, layer)
                for relation_row in _relation_rows(ferroelectric, layer):
                    if not include_non_switching and relation_row.get("reverses_P") is not True:
                        continue
                    rows.append(
                        {
                            "case_id": record.get("case_id"),
                            "file_name": record.get("file_name"),
                            "index": _case_value(record, result, "index", "summary.index"),
                            "conf": _case_value(record, result, "conf", "summary.conf"),
                            "phase": _case_value(record, result, "phase", "summary.phase"),
                            "sg_num": _case_value(
                                record,
                                result,
                                "input_space_group_number",
                                "groups.sg.number",
                                "group_identifiers.sg_num",
                            ),
                            "ossg_space_group_number": _case_value(
                                record,
                                result,
                                "ossg_space_group_number",
                                "groups.ossg.number",
                                "group_identifiers.ossg_space_group_number",
                            ),
                            "msg_parent_space_group_number": _case_value(
                                record,
                                result,
                                "msg_parent_space_group_number",
                                "groups.msg.parent_space_group_number",
                                "group_identifiers.msg_parent_space_group_number",
                            ),
                            "msg_symbol": _case_value(
                                record,
                                result,
                                "msg_symbol",
                                "groups.msg.symbol",
                                "group_identifiers.msg_symbol",
                            ),
                            "relation_layer": layer,
                            "screening_status": screening.get("status"),
                            "basis_setting": screening.get("basis_setting"),
                            "ordered_subgroup_source": screening.get("ordered_subgroup_source"),
                            "ordered_time_branch_scope": screening.get("ordered_time_branch_scope"),
                            "parent_space_group_number": screening.get("parent_space_group_number"),
                            "ordered_space_group_number": screening.get("ordered_space_group_number"),
                            "parent_grey_operation_count": screening.get("parent_grey_operation_count"),
                            "ordered_operation_count": screening.get("ordered_operation_count"),
                            "left_coset_count": screening.get("left_coset_count"),
                            "coset_index": relation_row.get("coset_index"),
                            "scope": relation_row.get("scope"),
                            "xyzt": relation_row.get("xyzt"),
                            "uvw": relation_row.get("uvw"),
                            "coset_operation": relation_row.get("coset_operation"),
                            "reverses_S": relation_row.get("reverses_S"),
                            "reverses_P": relation_row.get("reverses_P"),
                            "S_relation": relation_row.get("S_relation"),
                            "P_relation": relation_row.get("P_relation"),
                            "representative_class": relation_row.get("representative_class"),
                            "soc_allowed": relation_row.get("soc_allowed"),
                            "exchange_only": relation_row.get("exchange_only"),
                            "magnetic_order_relation": relation_row.get("magnetic_order_relation"),
                            "pattern_status": relation_row.get("pattern_status"),
                        }
                    )
    rows.sort(
        key=lambda row: (
            str(row.get("case_id") or ""),
            str(row.get("relation_layer") or ""),
            int(row.get("coset_index") or -1),
            str(row.get("xyzt") or ""),
            str(row.get("uvw") or ""),
        )
    )
    return rows


def _write_jsonl(rows: list[dict[str, Any]], output_jsonl: Path) -> None:
    output_jsonl.parent.mkdir(parents=True, exist_ok=True)
    with output_jsonl.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")


def _write_csv(rows: list[dict[str, Any]], output_csv: Path) -> None:
    output_csv.parent.mkdir(parents=True, exist_ok=True)
    with output_csv.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=ROW_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def _write_xlsx(rows: list[dict[str, Any]], output_xlsx: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "switching_ops"
    worksheet.append(ROW_COLUMNS)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        worksheet.append([row.get(column) for column in ROW_COLUMNS])
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for index, column in enumerate(ROW_COLUMNS, start=1):
        values = [column] + [str(row.get(column, "")) for row in rows[:200]]
        width = min(max(len(value) for value in values) + 2, 80)
        worksheet.column_dimensions[get_column_letter(index)].width = width
    workbook.save(output_xlsx)


def _summary(rows: list[dict[str, Any]]) -> dict[str, Any]:
    layer_counts = Counter(str(row.get("relation_layer")) for row in rows)
    case_counts = Counter(str(row.get("case_id")) for row in rows)
    relation_counts = Counter(
        (
            str(row.get("relation_layer")),
            "P+S" if row.get("reverses_S") is True else "P_only",
            str(row.get("representative_class")),
        )
        for row in rows
    )
    return {
        "row_count": len(rows),
        "case_count": len(case_counts),
        "layer_counts": dict(layer_counts),
        "relation_counts": {
            "|".join(key): count for key, count in sorted(relation_counts.items())
        },
    }


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Export ferroelectric-switching operation candidates from a "
            "FindSpinGroup full_results.jsonl file."
        )
    )
    parser.add_argument("--runtime-jsonl", type=Path, required=True)
    parser.add_argument("--output-jsonl", type=Path)
    parser.add_argument("--output-csv", type=Path)
    parser.add_argument("--output-xlsx", type=Path)
    parser.add_argument("--output-summary", type=Path)
    parser.add_argument(
        "--include-non-switching",
        action="store_true",
        help="Include rows that do not reverse P; by default only reverses_P=true rows are exported.",
    )
    args = parser.parse_args()

    rows = iter_switching_operation_rows(
        args.runtime_jsonl,
        include_non_switching=args.include_non_switching,
    )
    if args.output_jsonl is not None:
        _write_jsonl(rows, args.output_jsonl)
        print(args.output_jsonl)
    if args.output_csv is not None:
        _write_csv(rows, args.output_csv)
        print(args.output_csv)
    if args.output_xlsx is not None:
        _write_xlsx(rows, args.output_xlsx)
        print(args.output_xlsx)
    if args.output_summary is not None:
        args.output_summary.parent.mkdir(parents=True, exist_ok=True)
        args.output_summary.write_text(
            json.dumps(_summary(rows), indent=2, ensure_ascii=False),
            encoding="utf-8",
        )
        print(args.output_summary)
    if not any(
        item is not None
        for item in (
            args.output_jsonl,
            args.output_csv,
            args.output_xlsx,
            args.output_summary,
        )
    ):
        print(json.dumps(_summary(rows), indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
