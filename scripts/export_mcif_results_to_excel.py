from __future__ import annotations

import argparse
import json
import time
from pathlib import Path
from typing import Any

from findspingroup import batch_mcif
from findspingroup.find_spin_group import find_spin_group
from findspingroup.output_schema import (
    EXPORT_ROW_COLUMNS,
    MAGNETIC_ORBIT_EXPORT_COLUMNS,
    QUASI2D_EXPORT_COLUMNS,
    complete_export_row,
)
from findspingroup.structure.group import SpinSpaceGroup
from findspingroup.version import __version__ as FSG_VERSION


def _stringify(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _compact_wp_chain(wp_chain: Any, *, limit: int = 24) -> str | None:
    if not wp_chain:
        return None
    items: list[str] = []
    for row in wp_chain[:limit]:
        try:
            element, wp_sg, _idx_sg, wp_ssg, _idx_ssg, wp_msg, _idx_msg = row
            items.append(f"{element}:{wp_sg}->{wp_ssg}->{wp_msg}")
        except Exception:
            items.append(str(row))
    if len(wp_chain) > limit:
        items.append(f"...(+{len(wp_chain) - limit} more)")
    return " | ".join(items)


def _compact_magnetic_wp_dof_rows(rows: Any, *, limit: int = 16) -> str | None:
    if not rows:
        return None
    items: list[str] = []
    for row in list(rows)[:limit]:
        if not isinstance(row, dict):
            items.append(str(row))
            continue
        element = row.get("element")
        sg_wp = row.get("sg_wyckoff")
        ssg_wp = row.get("ssg_wyckoff_with_dof") or row.get("ssg_wyckoff")
        msg_wp = row.get("msg_wyckoff_with_dof") or row.get("msg_wyckoff")
        site_count = row.get("site_count")
        items.append(f"{element}:{sg_wp}->{ssg_wp}->{msg_wp} n={site_count}")
    if len(rows) > limit:
        items.append(f"...(+{len(rows) - limit} more)")
    return " | ".join(items)


def _normalize_spin_texture_config(config: Any) -> Any:
    if not isinstance(config, dict):
        return config
    normalized = dict(config)
    legacy_type_key = "wave" + "_type"
    if legacy_type_key in normalized and "spin_texture_type" not in normalized:
        normalized["spin_texture_type"] = normalized.pop(legacy_type_key)
    return normalized


def _spin_texture_type(config: Any) -> Any:
    if not isinstance(config, dict):
        return None
    return config.get("spin_texture_type", config.get("wave" + "_type"))


def _momentum_space_spin_configuration(config: Any) -> Any:
    if not isinstance(config, dict):
        return None
    return config.get("momentum_space_spin_configuration")


def _is_zero_expression(value: Any) -> bool:
    text = str(value).strip()
    if text in {"0", "0.0", "+0", "-0"}:
        return True
    try:
        return abs(float(text)) < 1e-12
    except ValueError:
        return False


def _component_equations(components: Any, *, symbol: str) -> list[str]:
    axes = ("x", "y", "z")
    equations: list[str] = []

    def visit(value: Any, indices: tuple[int, ...]) -> None:
        if isinstance(value, list):
            for index, item in enumerate(value):
                visit(item, (*indices, index))
            return
        if _is_zero_expression(value):
            return
        suffix = "".join(axes[index] if index < len(axes) else str(index) for index in indices)
        equations.append(f"{symbol}_{{{suffix}}} = {value}")

    visit(components, ())
    return equations


def _tensor_equations(payload: Any, *, symbol: str) -> str | None:
    if not isinstance(payload, dict):
        return None
    if "error" in payload:
        return f"ERROR: {payload['error']}"
    if payload.get("is_zero") is True:
        return f"{symbol}=0"

    parts: list[str] = []
    relations = payload.get("relations")
    if isinstance(relations, list) and relations:
        parts.append("; ".join(str(relation) for relation in relations))
    components = _component_equations(payload.get("components"), symbol=symbol)
    if components:
        parts.append("components: " + "; ".join(components))
    return " | ".join(parts) if parts else None


def _tensor_equation_export_values(tensor_outputs: Any) -> dict[str, Any]:
    payload = tensor_outputs if isinstance(tensor_outputs, dict) else {}
    return {
        "ahe_tensor_equations_soc": _tensor_equations(payload.get("AHE_wSOC"), symbol=r"\sigma"),
        "ahe_tensor_equations_no_soc": _tensor_equations(payload.get("AHE_woSOC"), symbol=r"\sigma"),
        "qmd_tensor_equations_soc": _tensor_equations(payload.get("MSGQMDTensor"), symbol="Q"),
        "qmd_tensor_equations_no_soc": _tensor_equations(payload.get("QMDTensor"), symbol="Q"),
        "imd_tensor_equations_soc": _tensor_equations(payload.get("MSGIMDTensor"), symbol="I"),
        "imd_tensor_equations_no_soc": _tensor_equations(payload.get("IMDTensor"), symbol="I"),
        "bcd_tensor_equations_soc": _tensor_equations(payload.get("MSGBCDTensor"), symbol="D"),
        "bcd_tensor_equations_no_soc": _tensor_equations(payload.get("BCDTensor"), symbol="D"),
    }


def _vector_axis_label(axis: Any) -> str | None:
    if not isinstance(axis, dict):
        return str(axis) if axis is not None else None
    label = axis.get("label")
    if label:
        return str(label)
    components = axis.get("components")
    if components is not None:
        return str(components)
    return None


def _axis_constraint_export_value(entry: Any, constraint_key: str) -> tuple[str | None, str | None]:
    constraint = {}
    if isinstance(entry, dict):
        constraints = entry.get("constraints")
        if isinstance(constraints, dict) and isinstance(constraints.get(constraint_key), dict):
            constraint = constraints[constraint_key]
    axes = constraint.get("allowed_axes") if isinstance(constraint, dict) else None
    labels = [
        label
        for axis in (axes or [])
        if (label := _vector_axis_label(axis)) is not None
    ]
    return (
        ", ".join(labels) if labels else None,
        constraint.get("allowed_axes_setting") if isinstance(constraint, dict) else None,
    )


def _vector_constraints_export_values(payload: Any) -> dict[str, Any]:
    constraints_payload = payload if isinstance(payload, dict) else {}
    values: dict[str, Any] = {}
    for key in ("sg", "ossg", "msg"):
        entry = constraints_payload.get(key) if isinstance(constraints_payload.get(key), dict) else {}
        polar_axes, polar_setting = _axis_constraint_export_value(
            entry,
            "real_space_t_even_p_odd",
        )
        axial_axes, axial_setting = _axis_constraint_export_value(
            entry,
            "real_space_t_even_p_even",
        )
        values[f"{key}_polar_axes"] = polar_axes
        values[f"{key}_polar_axes_setting"] = polar_setting
        values[f"{key}_real_space_axial_axes"] = axial_axes
        values[f"{key}_real_space_axial_axes_setting"] = axial_setting
    return values


def _has_wyckoff_splitting(
    wp_chain: Any,
    *,
    from_wp_index: int,
    from_orbit_index: int,
    to_wp_index: int,
    to_orbit_index: int,
) -> bool | None:
    if not wp_chain:
        return None
    mapping: dict[tuple[Any, Any, Any], set[tuple[Any, Any, Any]]] = {}
    saw_valid = False
    for row in wp_chain:
        if not isinstance(row, (list, tuple)) or len(row) <= max(to_wp_index, to_orbit_index):
            continue
        element = row[0]
        source = (element, row[from_wp_index], row[from_orbit_index])
        target = (element, row[to_wp_index], row[to_orbit_index])
        if source[1] in (None, "") or target[1] in (None, ""):
            continue
        saw_valid = True
        mapping.setdefault(source, set()).add(target)
    if not saw_valid:
        return None
    return any(len(targets) > 1 for targets in mapping.values())


def _wyckoff_splitting_export_values(wp_chain: Any) -> dict[str, Any]:
    return {
        "has_wyckoff_splitting_sg_to_ossg": _has_wyckoff_splitting(
            wp_chain,
            from_wp_index=1,
            from_orbit_index=2,
            to_wp_index=3,
            to_orbit_index=4,
        ),
        "has_wyckoff_splitting_ossg_to_msg": _has_wyckoff_splitting(
            wp_chain,
            from_wp_index=3,
            from_orbit_index=4,
            to_wp_index=5,
            to_orbit_index=6,
        ),
    }


def _magnetic_site_export_values(summary: Any) -> dict[str, Any]:
    payload = summary if isinstance(summary, dict) else {}
    magnetic_wp_dof_rows = payload.get("magnetic_wp_dof_rows")
    return {
        "magnetic_site_setting": payload.get("setting"),
        "magnetic_site_sg_primitive_to_magnetic_primitive_cell_expansion": (
            payload.get("cell_expansion")
        ),
        "magnetic_atom_count": payload.get("magnetic_atom_count"),
        "nonzero_moment_atom_count": payload.get("nonzero_moment_atom_count"),
        "zero_moment_magnetic_atom_count": payload.get("zero_moment_magnetic_atom_count"),
        "number_of_magnetic_orbits_sg": payload.get("n_magnetic_orbits_sg"),
        "number_of_magnetic_orbits_ssg": payload.get("n_magnetic_orbits_ssg"),
        "number_of_magnetic_orbits_msg": payload.get("n_magnetic_orbits_msg"),
        "max_magnetic_site_dof_ssg": payload.get("max_magnetic_site_dof_ssg"),
        "max_magnetic_site_dof_msg": payload.get("max_magnetic_site_dof_msg"),
        "total_magnetic_site_dof_ssg": payload.get("total_magnetic_site_dof_ssg"),
        "total_magnetic_site_dof_msg": payload.get("total_magnetic_site_dof_msg"),
        "magnetic_wyckoff_dof_summary": _compact_magnetic_wp_dof_rows(magnetic_wp_dof_rows),
        "_magnetic_site_orbit_rows": magnetic_wp_dof_rows,
    }


def _quasi2d_export_values(quasi_2d: Any) -> dict[str, Any]:
    if not isinstance(quasi_2d, dict):
        return {}
    payload = quasi_2d
    no_soc_config = payload.get("spin_texture_config_no_soc", payload.get("wave" + "_spin_config_no_soc"))
    soc_config = payload.get("spin_texture_config_soc", payload.get("wave" + "_spin_config_soc"))
    diagnostic_points = payload.get("diagnostic_points") or []
    generated_point = diagnostic_points[0] if diagnostic_points else {}
    kpoints = payload.get("kpoints") or []
    projection_summary = payload.get("kpoint_projection_summary") or {}
    return {
        "quasi2d_status": payload.get("status"),
        "quasi2d_source": payload.get("source"),
        "vacuum_axis_input": payload.get("vacuum_axis_input"),
        "spin_splitting_2d": payload.get("spin_splitting_2d"),
        "spin_splitting_2d_interpretation": payload.get("interpretation"),
        "is_alter_2d": payload.get("is_alter_2d"),
        "quasi2d_magnetic_phase": payload.get("magnetic_phase"),
        "quasi2d_gp_label": generated_point.get("label"),
        "quasi2d_gp_symbol": generated_point.get("k_symbol_2d"),
        "quasi2d_gp_k_input": generated_point.get("k_input_reciprocal"),
        "quasi2d_gp_k_acc": generated_point.get("k_acc_primitive"),
        "quasi2d_gp_spin_splitting": generated_point.get("spin_splitting"),
        "quasi2d_gp_spin_polarizations": generated_point.get("spin_polarizations"),
        "quasi2d_spin_texture_config_no_soc": _normalize_spin_texture_config(no_soc_config),
        "quasi2d_spin_texture_type_no_soc": _spin_texture_type(no_soc_config),
        "quasi2d_momentum_space_spin_configuration_no_soc": (
            _momentum_space_spin_configuration(no_soc_config)
        ),
        "quasi2d_spin_texture_config_soc": _normalize_spin_texture_config(soc_config),
        "quasi2d_spin_texture_type_soc": _spin_texture_type(soc_config),
        "quasi2d_momentum_space_spin_configuration_soc": (
            _momentum_space_spin_configuration(soc_config)
        ),
        "quasi2d_spin_texture_basis": payload.get(
            "spin_texture_config_basis",
            payload.get("wave" + "_spin_config_basis"),
        ),
        "quasi2d_kpoint_projection_summary": projection_summary,
        "quasi2d_kpoints": [
            {
                "label": row.get("label"),
                "k_symbol_2d": row.get("k_symbol_2d"),
                "plane": row.get("plane_classification"),
                "spin_splitting": row.get("spin_splitting"),
            }
            for row in kpoints
        ],
    }


def _serialized_property(payload: dict[str, Any], payload_key: str, property_key: str) -> Any:
    properties = payload.get("properties") if isinstance(payload.get("properties"), dict) else {}
    if payload_key in payload:
        return payload.get(payload_key)
    return properties.get(property_key)


def _serialized_payload_value(payload: dict[str, Any], key: str, *legacy_keys: str) -> Any:
    if key in payload:
        return payload.get(key)
    for legacy_key in legacy_keys:
        if legacy_key in payload:
            return payload.get(legacy_key)
    return None


def _row_from_result(file_path: Path, result, *, duration_seconds: float | None = None) -> dict[str, Any]:
    identify = result.identify_index_details or {}
    primitive_ssg = SpinSpaceGroup(result.primitive_magnetic_cell_ssg_ops)
    no_soc_config = result.spin_texture_config_no_soc
    soc_config = result.spin_texture_config_soc

    row = {
        "case_id": batch_mcif._normalize_case_id(file_path),
        "file_name": file_path.name,
        "duration_seconds": duration_seconds,
        "index": result.index,
        "conf": result.conf,
        "phase": result.magnetic_phase,
        "acc": result.acc,
        "msg_acc": result.msg_acc,
        "G0_id": identify.get("G0_id"),
        "L0_id": identify.get("L0_id"),
        "t_index": identify.get("t_index"),
        "k_index": identify.get("k_index"),
        "nsspg_hm": primitive_ssg.n_spin_part_point_group_symbol_hm,
        "nsspg_symbol": primitive_ssg.n_spin_part_point_group_symbol_s,
        "sspg_hm": primitive_ssg.spin_part_point_group_symbol_hm,
        "sspg_symbol": primitive_ssg.spin_part_point_group_symbol_s,
        "ssg_type": result.primitive_magnetic_cell_ssg_type,
        "spin_only_direction(ossg convention)": result.convention_spin_only_direction,
        "ossg_symbol_linear": result.convention_ssg_international_linear,
        "ossg_symbol_latex": result.convention_ssg_international_latex,
        "sg_symbol": result.input_space_group_symbol,
        "sg_num": result.input_space_group_number,
        "sg_is_centrosymmetric": result.sg_is_centrosymmetric,
        "sg_is_polar": result.sg_is_polar,
        "sg_is_chiral": result.sg_is_chiral,
        "ossg_space_group_number": result.ossg_space_group_number,
        "ossg_is_centrosymmetric": result.ossg_is_centrosymmetric,
        "ossg_is_polar": result.ossg_is_polar,
        "ossg_is_chiral": result.ossg_is_chiral,
        "msg_symbol": result.msg_symbol,
        "msg_num": result.msg_num,
        "msg_type": result.msg_type,
        "msg_bns_number": result.msg_bns_number,
        "msg_og_number": result.msg_og_number,
        "msg_parent_space_group_number": result.msg_parent_space_group_number,
        "msg_is_centrosymmetric": result.msg_is_centrosymmetric,
        "msg_is_polar": result.msg_is_polar,
        "msg_is_chiral": result.msg_is_chiral,
        "empg_symbol": result.gspg_effective_mpg_symbol,
        "spin_splitting_with_soc": result.spinsplitting_w_soc,
        "spin_splitting_without_soc": result.spinsplitting_wo_soc,
        "spin_texture_config_soc": _normalize_spin_texture_config(soc_config),
        "spin_texture_type_soc": _spin_texture_type(soc_config),
        "momentum_space_spin_configuration_soc": (
            _momentum_space_spin_configuration(soc_config)
        ),
        "spin_texture_config_no_soc": _normalize_spin_texture_config(no_soc_config),
        "spin_texture_type_no_soc": _spin_texture_type(no_soc_config),
        "momentum_space_spin_configuration_no_soc": (
            _momentum_space_spin_configuration(no_soc_config)
        ),
        "ahc_with_soc": result.ahc_w_soc,
        "ahc_without_soc": result.ahc_wo_soc,
        "is_altermagnet": result.is_alter,
        "is_spin_orbit_magnet": result.is_spin_orbit_magnet,
        "wyckoff_split": _compact_wp_chain(result.wp_chain),
        "acc_primitive_wyckoff_split": _compact_wp_chain(result.acc_primitive_wp_chain),
    }
    row.update(_tensor_equation_export_values(getattr(result, "tensor_outputs", None)))
    row.update(
        _vector_constraints_export_values(
            getattr(result, "vector_constraints_by_symmetry", None)
        )
    )
    row.update(_wyckoff_splitting_export_values(result.wp_chain))
    row.update(_magnetic_site_export_values(getattr(result, "magnetic_site_summary", None)))
    row.update(_quasi2d_export_values(getattr(result, "quasi_2d", None)))
    return complete_export_row(row)


def _row_from_serialized_result_record(record: dict[str, Any]) -> dict[str, Any]:
    payload = record.get("result") or {}
    identify = payload.get("identify_index_details") or {}
    primitive_ops = payload.get("primitive_magnetic_cell_ssg_ops") or []
    primitive_ssg = SpinSpaceGroup(primitive_ops) if primitive_ops else None
    no_soc_config = payload.get("spin_texture_config_no_soc", payload.get("wave" + "_spin_config_no_soc"))
    soc_config = payload.get("spin_texture_config_soc", payload.get("wave" + "_spin_config_soc"))

    row = {
        "case_id": record.get("case_id"),
        "file_name": record.get("file_name"),
        "duration_seconds": record.get("duration_seconds"),
        "index": payload.get("index"),
        "conf": payload.get("conf"),
        "phase": payload.get("phase"),
        "acc": payload.get("acc"),
        "msg_acc": payload.get("msg_acc"),
        "G0_id": identify.get("G0_id"),
        "L0_id": identify.get("L0_id"),
        "t_index": identify.get("t_index"),
        "k_index": identify.get("k_index"),
        "nsspg_hm": (
            primitive_ssg.n_spin_part_point_group_symbol_hm if primitive_ssg is not None else None
        ),
        "nsspg_symbol": (
            primitive_ssg.n_spin_part_point_group_symbol_s if primitive_ssg is not None else None
        ),
        "sspg_hm": primitive_ssg.spin_part_point_group_symbol_hm if primitive_ssg is not None else None,
        "sspg_symbol": primitive_ssg.spin_part_point_group_symbol_s if primitive_ssg is not None else None,
        "ssg_type": payload.get("primitive_magnetic_cell_ssg_type"),
        "spin_only_direction(ossg convention)": payload.get("convention_spin_only_direction"),
        "ossg_symbol_linear": payload.get("convention_ssg_international_linear"),
        "ossg_symbol_latex": payload.get("convention_ssg_international_latex"),
        "sg_symbol": payload.get("input_space_group_symbol"),
        "sg_num": payload.get("input_space_group_number"),
        "sg_is_centrosymmetric": _serialized_payload_value(
            payload,
            "sg_is_centrosymmetric",
            "sg_has_real_space_inversion",
        ),
        "sg_is_polar": payload.get("sg_is_polar"),
        "sg_is_chiral": payload.get("sg_is_chiral"),
        "ossg_space_group_number": payload.get("ossg_space_group_number"),
        "ossg_is_centrosymmetric": _serialized_payload_value(
            payload,
            "ossg_is_centrosymmetric",
            "ossg_has_real_space_inversion",
        ),
        "ossg_is_polar": payload.get("ossg_is_polar"),
        "ossg_is_chiral": payload.get("ossg_is_chiral"),
        "msg_symbol": payload.get("msg_symbol"),
        "msg_num": payload.get("msg_num"),
        "msg_type": payload.get("msg_type"),
        "msg_bns_number": payload.get("msg_bns_number"),
        "msg_og_number": payload.get("msg_og_number"),
        "msg_parent_space_group_number": payload.get("msg_parent_space_group_number"),
        "msg_is_centrosymmetric": _serialized_payload_value(
            payload,
            "msg_is_centrosymmetric",
            "msg_has_real_space_inversion",
        ),
        "msg_is_polar": payload.get("msg_is_polar"),
        "msg_is_chiral": payload.get("msg_is_chiral"),
        "empg_symbol": payload.get("gspg_effective_mpg_symbol"),
        "spin_splitting_with_soc": _serialized_property(payload, "spinsplitting_w_soc", "ss_w_soc"),
        "spin_splitting_without_soc": _serialized_property(payload, "spinsplitting_wo_soc", "ss_wo_soc"),
        "spin_texture_config_soc": _normalize_spin_texture_config(soc_config),
        "spin_texture_type_soc": _spin_texture_type(soc_config),
        "momentum_space_spin_configuration_soc": (
            _momentum_space_spin_configuration(soc_config)
        ),
        "spin_texture_config_no_soc": _normalize_spin_texture_config(no_soc_config),
        "spin_texture_type_no_soc": _spin_texture_type(no_soc_config),
        "momentum_space_spin_configuration_no_soc": (
            _momentum_space_spin_configuration(no_soc_config)
        ),
        "ahc_with_soc": _serialized_property(payload, "ahc_w_soc", "ahc_w_soc"),
        "ahc_without_soc": _serialized_property(payload, "ahc_wo_soc", "ahc_wo_soc"),
        "is_altermagnet": _serialized_property(payload, "is_alter", "is_alter"),
        "is_spin_orbit_magnet": _serialized_property(
            payload,
            "is_spin_orbit_magnet",
            "is_spin_orbit_magnet",
        ),
        "wyckoff_split": _compact_wp_chain(payload.get("wp_chain")),
        "acc_primitive_wyckoff_split": _compact_wp_chain(
            payload.get("acc_primitive_wp_chain")
        ),
        "error_type": record.get("error", {}).get("type"),
        "error_message": record.get("error", {}).get("message"),
    }
    row.update(_tensor_equation_export_values(payload.get("tensor_outputs")))
    row.update(
        _vector_constraints_export_values(payload.get("vector_constraints_by_symmetry"))
    )
    row.update(_wyckoff_splitting_export_values(payload.get("wp_chain")))
    row.update(_magnetic_site_export_values(payload.get("magnetic_site_summary")))
    row.update(_quasi2d_export_values(payload.get("quasi_2d")))
    return complete_export_row(row)


def _row_from_error(
    file_path: Path,
    exc: Exception,
    *,
    duration_seconds: float | None = None,
) -> dict[str, Any]:
    return complete_export_row({
        "case_id": batch_mcif._normalize_case_id(file_path),
        "file_name": file_path.name,
        "duration_seconds": duration_seconds,
        "error_type": type(exc).__name__,
        "error_message": str(exc),
    })


def _runtime_export_metadata(runtime_jsonl: Path | None) -> dict[str, Any]:
    if runtime_jsonl is None:
        return {
            "source_fsg_version": FSG_VERSION,
            "source_run_tag": None,
        }

    summary_path = runtime_jsonl.parent / "summary.json"
    metadata = {
        "source_fsg_version": None,
        "source_run_tag": runtime_jsonl.parent.name,
    }
    if summary_path.exists():
        try:
            summary = json.loads(summary_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            summary = {}
        metadata.update(
            {
                "source_fsg_version": summary.get("package_version"),
                "source_run_tag": summary.get("run_tag") or runtime_jsonl.parent.name,
            }
        )
    return metadata


def _runtime_export_metadata_from_paths(runtime_jsonls: list[Path]) -> dict[str, Any]:
    if not runtime_jsonls:
        return {
            "source_fsg_version": FSG_VERSION,
            "source_run_tag": None,
        }
    metadata = _runtime_export_metadata(runtime_jsonls[0])
    run_tags = []
    versions = set()
    for runtime_jsonl in runtime_jsonls:
        item = _runtime_export_metadata(runtime_jsonl)
        if item.get("source_run_tag"):
            run_tags.append(item["source_run_tag"])
        if item.get("source_fsg_version"):
            versions.add(item["source_fsg_version"])
    if len(runtime_jsonls) > 1:
        parent = runtime_jsonls[0].parent.parent
        metadata["source_run_tag"] = parent.name if parent.name else "sharded_runtime"
    if len(versions) == 1:
        metadata["source_fsg_version"] = next(iter(versions))
    return metadata


def _apply_export_metadata(rows: list[dict[str, Any]], metadata: dict[str, Any]) -> None:
    for row in rows:
        row["source_fsg_version"] = metadata.get("source_fsg_version")
        row["source_run_tag"] = metadata.get("source_run_tag")


COLUMNS = list(EXPORT_ROW_COLUMNS)


MAGNETIC_ORBIT_COLUMNS = list(MAGNETIC_ORBIT_EXPORT_COLUMNS)


QUASI2D_COLUMNS = list(QUASI2D_EXPORT_COLUMNS)


QUASI2D_RECORD_COLUMNS = [
    column for column in QUASI2D_COLUMNS if column not in COLUMNS
]

TEXT_FORMAT_COLUMNS = {
    "index",
    "msg_num",
    "msg_bns_number",
    "msg_og_number",
}


def _xlsx_cell_value(column: str, value: Any) -> Any:
    if value is None:
        return None
    if column in TEXT_FORMAT_COLUMNS:
        return str(value)
    return _stringify(value)


def _format_text_columns(ws, columns: list[str]) -> None:
    text_column_indices = [
        column_index
        for column_index, column in enumerate(columns, start=1)
        if column in TEXT_FORMAT_COLUMNS
    ]
    for column_index in text_column_indices:
        for row in ws.iter_rows(min_row=2, min_col=column_index, max_col=column_index):
            cell = row[0]
            if cell.value is None:
                continue
            cell.value = str(cell.value)
            cell.number_format = "@"


def _has_quasi2d_values(row: dict[str, Any]) -> bool:
    return any(row.get(column) is not None for column in QUASI2D_RECORD_COLUMNS)


def _record_columns_for_rows(rows: list[dict[str, Any]]) -> list[str]:
    if any(_has_quasi2d_values(row) for row in rows):
        return [*COLUMNS, *QUASI2D_RECORD_COLUMNS]
    return COLUMNS


def _write_workbook(rows: list[dict[str, Any]], output_xlsx: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "records"
    record_columns = _record_columns_for_rows(rows)
    ws.append(record_columns)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([_xlsx_cell_value(column, row.get(column)) for column in record_columns])
    _format_text_columns(ws, record_columns)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for index, column in enumerate(record_columns, start=1):
        max_len = len(column)
        for row in ws.iter_rows(min_row=2, min_col=index, max_col=index):
            value = row[0].value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(index)].width = min(max_len + 2, 60)

    orbit_ws = wb.create_sheet("magnetic_site_orbits")
    orbit_ws.append(MAGNETIC_ORBIT_COLUMNS)
    for cell in orbit_ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        for orbit in row.get("_magnetic_site_orbit_rows") or []:
            if not isinstance(orbit, dict):
                continue
            orbit_row = {
                "case_id": row.get("case_id"),
                "file_name": row.get("file_name"),
                "index": row.get("index"),
                **orbit,
            }
            orbit_ws.append(
                [
                    _xlsx_cell_value(column, orbit_row.get(column))
                    for column in MAGNETIC_ORBIT_COLUMNS
                ]
            )
    _format_text_columns(orbit_ws, MAGNETIC_ORBIT_COLUMNS)

    orbit_ws.freeze_panes = "A2"
    orbit_ws.auto_filter.ref = orbit_ws.dimensions
    for index, column in enumerate(MAGNETIC_ORBIT_COLUMNS, start=1):
        max_len = len(column)
        for row in orbit_ws.iter_rows(min_row=2, min_col=index, max_col=index):
            value = row[0].value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        orbit_ws.column_dimensions[get_column_letter(index)].width = min(max_len + 2, 60)

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)


def _write_jsonl(rows: list[dict[str, Any]], output_jsonl: Path) -> None:
    output_jsonl.parent.mkdir(parents=True, exist_ok=True)
    with output_jsonl.open("w", encoding="utf-8") as handle:
        for row in rows:
            public_row = {key: value for key, value in row.items() if not key.startswith("_")}
            handle.write(json.dumps(public_row, ensure_ascii=False) + "\n")


def _discover_runtime_jsonls(
    runtime_jsonls: list[Path] | None,
    *,
    shard_root: Path | None,
    shard_glob: str,
) -> list[Path]:
    paths = list(runtime_jsonls or [])
    if shard_root is not None:
        shard_dirs = [
            path
            for path in shard_root.glob(shard_glob)
            if path.is_dir() and (path / "full_results.jsonl").exists()
        ]

        def sort_key(path: Path) -> tuple[int, str]:
            try:
                return (int(path.name.rsplit("_", 1)[1]), path.name)
            except (IndexError, ValueError):
                return (10**9, path.name)

        paths.extend(path / "full_results.jsonl" for path in sorted(shard_dirs, key=sort_key))
    seen: set[Path] = set()
    unique_paths: list[Path] = []
    for path in paths:
        resolved = path.resolve()
        if resolved in seen:
            continue
        seen.add(resolved)
        unique_paths.append(path)
    return unique_paths


def _rows_from_runtime_jsonls(
    runtime_jsonls: list[Path],
    *,
    limit: int | None = None,
    sort_rows: bool = False,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    count = 0
    total = limit if limit is not None else "?"
    for runtime_jsonl in runtime_jsonls:
        with runtime_jsonl.open(encoding="utf-8") as handle:
            for line in handle:
                if not line.strip():
                    continue
                if limit is not None and count >= limit:
                    break
                record = json.loads(line)
                count += 1
                rows.append(_row_from_serialized_result_record(record))
                print(
                    f"[{count}/{total}] {record.get('status', 'ok').upper():5} "
                    f"{record.get('file_name')} ({runtime_jsonl.parent.name})"
                )
        if limit is not None and count >= limit:
            break
    if sort_rows:
        rows.sort(key=lambda row: (str(row.get("case_id") or ""), str(row.get("file_name") or "")))
    return rows


def main() -> None:
    parser = argparse.ArgumentParser(description="Run find_spin_group over mcif files and export compact Excel rows.")
    parser.add_argument("inputs", nargs="*", help="Input .mcif files or directories")
    parser.add_argument(
        "--runtime-jsonl",
        type=Path,
        nargs="+",
        help="Read rows from one or more batch full_results.jsonl files instead of re-running.",
    )
    parser.add_argument(
        "--shard-root",
        type=Path,
        help="Read shard_*/full_results.jsonl files from a parallel batch output root.",
    )
    parser.add_argument(
        "--shard-glob",
        default="shard_*",
        help="Glob used under --shard-root to discover shard directories.",
    )
    parser.add_argument("--output-xlsx", type=Path)
    parser.add_argument("--output-jsonl", type=Path)
    parser.add_argument("--limit", type=int)
    parser.add_argument("--non-recursive", action="store_true")
    parser.add_argument("--space-tol", type=float, default=0.02)
    parser.add_argument("--mtol", type=float, default=0.02)
    parser.add_argument("--meigtol", type=float, default=0.00002)
    parser.add_argument("--matrix-tol", type=float, default=0.01)
    args = parser.parse_args()
    if args.output_xlsx is None and args.output_jsonl is None:
        raise ValueError("Provide at least one of --output-xlsx or --output-jsonl.")
    rows: list[dict[str, Any]] = []
    runtime_jsonls = _discover_runtime_jsonls(
        args.runtime_jsonl,
        shard_root=args.shard_root,
        shard_glob=args.shard_glob,
    )
    export_metadata = _runtime_export_metadata_from_paths(runtime_jsonls)
    if runtime_jsonls:
        rows = _rows_from_runtime_jsonls(
            runtime_jsonls,
            limit=args.limit,
            sort_rows=args.shard_root is not None or len(runtime_jsonls) > 1,
        )
    else:
        if not args.inputs:
            raise ValueError(
                "Provide input files/directories unless --runtime-jsonl or --shard-root is used."
            )
        files = batch_mcif._discover_mcif_files(args.inputs, recursive=not args.non_recursive)
        files = batch_mcif._dedupe_sorted(files)
        if args.limit is not None:
            files = files[: args.limit]

        for index, file_path in enumerate(files, start=1):
            case_start = time.perf_counter()
            try:
                result = find_spin_group(
                    str(file_path),
                    space_tol=args.space_tol,
                    mtol=args.mtol,
                    meigtol=args.meigtol,
                    matrix_tol=args.matrix_tol,
                )
                duration = round(time.perf_counter() - case_start, 6)
                rows.append(_row_from_result(file_path, result, duration_seconds=duration))
                print(f"[{index}/{len(files)}] OK    {file_path.name} -> {result.index} ({duration:.3f}s)")
            except Exception as exc:
                duration = round(time.perf_counter() - case_start, 6)
                rows.append(_row_from_error(file_path, exc, duration_seconds=duration))
                print(
                    f"[{index}/{len(files)}] ERROR {file_path.name} -> "
                    f"{type(exc).__name__}: {exc} ({duration:.3f}s)"
                )

    _apply_export_metadata(rows, export_metadata)

    if args.output_jsonl is not None:
        _write_jsonl(rows, args.output_jsonl)
        print(args.output_jsonl)
    if args.output_xlsx is not None:
        _write_workbook(rows, args.output_xlsx)
        print(args.output_xlsx)


if __name__ == "__main__":
    main()
