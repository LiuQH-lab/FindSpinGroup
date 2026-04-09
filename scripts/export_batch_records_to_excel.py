from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


DEFAULT_COLUMNS = [
    ("case_id", "case_id"),
    ("file_name", "file_name"),
    ("status", "status"),
    ("duration_seconds", "duration_seconds"),
    ("result.index", "index"),
    ("result.conf", "conf"),
    ("result.phase", "phase"),
    ("result.acc", "acc"),
    ("result.msg_acc", "msg_acc"),
    ("result.gspg.symbol_linear", "gspg_symbol"),
    ("result.gspg.effective_mpg_symbol", "effective_mpg"),
    ("result.gspg.npg_symbol_s", "npg_symbol"),
    ("result.gspg.spin_only_component_symbol_s", "spin_only_component"),
    ("result.gspg.real_space_setting", "gspg_real_setting"),
    ("result.gspg.spin_frame_setting", "gspg_spin_frame"),
    ("result.properties.ss_w_soc", "ss_w_soc"),
    ("result.properties.ss_wo_soc", "ss_wo_soc"),
    ("result.properties.ahc_w_soc", "ahc_w_soc"),
    ("result.properties.ahc_wo_soc", "ahc_wo_soc"),
    ("tensor_summary.AHE_wSOC.free_parameters", "AHE_wSOC_free"),
    ("tensor_summary.AHE_wSOC.is_zero", "AHE_wSOC_zero"),
    ("tensor_summary.AHE_woSOC.free_parameters", "AHE_woSOC_free"),
    ("tensor_summary.AHE_woSOC.is_zero", "AHE_woSOC_zero"),
    ("tensor_summary.BCDTensor.free_parameters", "BCD_free"),
    ("tensor_summary.BCDTensor.is_zero", "BCD_zero"),
    ("tensor_summary.IMDTensor.free_parameters", "IMD_free"),
    ("tensor_summary.IMDTensor.is_zero", "IMD_zero"),
    ("tensor_summary.MSGBCDTensor.free_parameters", "MSG_BCD_free"),
    ("tensor_summary.MSGBCDTensor.is_zero", "MSG_BCD_zero"),
    ("tensor_summary.MSGIMDTensor.free_parameters", "MSG_IMD_free"),
    ("tensor_summary.MSGIMDTensor.is_zero", "MSG_IMD_zero"),
    ("tensor_summary.MSGQMDTensor.free_parameters", "MSG_QMD_free"),
    ("tensor_summary.MSGQMDTensor.is_zero", "MSG_QMD_zero"),
]


def _get_path(data: dict[str, Any], path: str) -> Any:
    current: Any = data
    for part in path.split("."):
        if not isinstance(current, dict) or part not in current:
            return None
        current = current[part]
    return current


def _normalize_value(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _load_records(records_path: Path) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    with records_path.open("r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            records.append(json.loads(line))
    return records


def _write_summary_sheet(workbook: Workbook, summary: dict[str, Any] | None, records_count: int) -> None:
    ws = workbook.active
    ws.title = "summary"
    rows = [
        ("records_count", records_count),
    ]
    if summary is not None:
        rows.extend(
            [
                ("output_dir", summary.get("output_dir")),
                ("package_version", summary.get("package_version")),
                ("started_at", summary.get("started_at")),
                ("finished_at", summary.get("finished_at")),
                ("processed_cases", summary.get("processed_cases")),
                ("success_count", summary.get("success_count")),
                ("error_count", summary.get("error_count")),
                ("space_tol", _get_path(summary, "tolerances.space_tol")),
                ("mtol", _get_path(summary, "tolerances.mtol")),
                ("meigtol", _get_path(summary, "tolerances.meigtol")),
                ("matrix_tol", _get_path(summary, "tolerances.matrix_tol")),
            ]
        )

    ws.append(["field", "value"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for key, value in rows:
        ws.append([key, _normalize_value(value)])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 80


def _write_records_sheet(workbook: Workbook, records: list[dict[str, Any]]) -> None:
    ws = workbook.create_sheet("records")
    ws.append([header for _, header in DEFAULT_COLUMNS])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for record in records:
        ws.append([_normalize_value(_get_path(record, path)) for path, _header in DEFAULT_COLUMNS])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for index, (_path, header) in enumerate(DEFAULT_COLUMNS, start=1):
        max_len = len(header)
        for row in ws.iter_rows(min_row=2, min_col=index, max_col=index):
            value = row[0].value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(index)].width = min(max_len + 2, 40)


def main() -> None:
    parser = argparse.ArgumentParser(description="Export compact batch records to Excel.")
    parser.add_argument("records_jsonl", type=Path, help="Path to records.jsonl")
    parser.add_argument("output_xlsx", type=Path, help="Path to output .xlsx")
    parser.add_argument("--summary-json", type=Path, help="Optional summary.json for a summary sheet")
    args = parser.parse_args()

    records = _load_records(args.records_jsonl)
    summary = None
    if args.summary_json is not None and args.summary_json.exists():
        summary = json.loads(args.summary_json.read_text(encoding="utf-8"))

    workbook = Workbook()
    _write_summary_sheet(workbook, summary, len(records))
    _write_records_sheet(workbook, records)

    args.output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output_xlsx)
    print(args.output_xlsx)


if __name__ == "__main__":
    main()
